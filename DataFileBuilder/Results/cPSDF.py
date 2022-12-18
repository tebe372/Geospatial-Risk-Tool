import numpy as np
from openpyxl import load_workbook
import copy
import time
from scipy import sparse
import matplotlib.pyplot as plt
from matplotlib.cm import get_cmap
from mpl_toolkits.basemap import Basemap
import sys
import os

# ############### subfunctions ###############
def loadData(filename):
    wb = load_workbook(filename)
    sheets = wb.worksheets
    sheet1 = sheets[0]
    nFiles = 0

    txt = []
    for col in sheet1['A']:
        txt.append(col.value)
        nFiles += 1

    num = []
    for col in sheet1['B']:
        num.append(col.value)

    nConcSamples = 0
    mDataCell = []
    tmp = [1,2,3]
    for j in range(nFiles):
        if num[j] >= -1e-64:
            nConcSamples = nConcSamples + 1
            tmp[0] = num[j]
            tmp[1] = 1
            tmp[2] = txt[j]
        else:
            tmp[0] = num[j]
            tmp[1] = tmp[1] + 1
            tmp[2] = txt[j]
        mDataCell.append(copy.deepcopy(tmp))

    return mDataCell, nConcSamples


def ReadDateFile(directory, filename):
    out = []
    fid = open(f'{directory}/{filename}', 'r')
    for i in range(1,10):
        fid.readline()
    while True:
        line = fid.readline().split()
        if not line:
            break
        out.append(list(map(float, line)))
    fid.close()
    return np.array(out)


def createWxu(xList, yList, xTrCent, yTrCent, dtCent, tCent, tau):
    Nx = len(xList)
    Ny = len(yList)
    Npoints = len(xTrCent)
    Wxu = np.zeros(Nx*Ny)
    if Npoints > 1:
        for j in range(Npoints):
            xx = xTrCent[j]
            yy = yTrCent[j]
            if tau == 0:
                tt = dtCent[j]
            else:
                tt=dtCent[j] * np.exp(tCent[j]/tau)
            if ( xx > min(xList) and xx < max(xList) and yy > min(yList) and yy < max(yList) ):
                idx1 = np.argwhere(xList <= xx)
                Nidx = len(idx1)
                iidx1 = idx1[Nidx-1]
                idx2 = np.argwhere(xList > xx)
                iidx2 = idx2[0]
                x1 = xList[iidx1]
                x2 = xList[iidx2]
                a1 = 1-(xx-x1)/(x2-x1)
                a2 = (xx-x1)/(x2-x1)

                idy1 = np.argwhere(yList <= yy)
                Nidy = len(idy1)
                iidy1 = idy1[Nidy-1]
                idy2 = np.argwhere(yList > yy)
                iidy2 = idy2[0]
                y1 = yList[iidy1]
                y2 = yList[iidy2]
                b1 = 1-(yy-y1)/(y2-y1)
                b2 = (yy-y1)/(y2-y1)

                Wxu[iidy1*Nx +iidx1] += a1*b1*tt
                Wxu[iidy1*Nx +iidx2] += a2*b1*tt
                Wxu[iidy2*Nx +iidx1] += a1*b2*tt
                Wxu[iidy2*Nx +iidx2] += a2*b2*tt
    return Wxu


def createWxuStar(xList, yList, xx, yy):
    Nx = len(xList)
    Ny = len(yList)

    Wxu = np.zeros(Nx*Ny)
    if ( xx > min(xList) and xx < max(xList) and yy > min(yList) and yy < max(yList) ):
        idx1 = np.argwhere(xList <= xx)
        Nidx = len(idx1)
        iidx1 = idx1[Nidx-1]
        idx2 = np.argwhere(xList > xx)
        iidx2 = idx2[0]
        x1 = xList[iidx1]
        x2 = xList[iidx2]
        a1 = 1-(xx-x1)/(x2-x1)
        a2 = (xx-x1)/(x2-x1)
            
        idy1 = np.argwhere(yList <= yy)
        Nidy = len(idy1)
        iidy1 = idy1[Nidy-1]
        idy2 = np.argwhere(yList > yy)
        iidy2 = idy2[0]
        y1 = yList[iidy1]
        y2 = yList[iidy2]
        b1 = 1-(yy-y1)/(y2-y1)
        b2 = (yy-y1)/(y2-y1)
        
        Wxu[iidy1*Nx +iidx1] += a1*b1
        Wxu[iidy1*Nx +iidx2] += a2*b1
        Wxu[iidy2*Nx +iidx1] += a1*b2
        Wxu[iidy2*Nx +iidx2] += a2*b2
    return Wxu


def bareSE1d(x1, x2, len):
    y = np.exp(-(x1-x2)**2/(2*len*len))
    return y


def createJuu(xList, yList, len):
    Nx = np.size(xList)
    xJuu1 = np.zeros(Nx)
    for j in range(Nx):
        ttt = bareSE1d(xList[0], xList[j], len)
        if ttt >= 1e-6:
            xJuu1[j] = ttt

    Ny = np.size(yList)
    yJuu1 = np.zeros(Ny)
    for j in range(Ny):
        ttt = bareSE1d(yList[0], yList[j], len)
        if ttt >= 1e-6:
            yJuu1[j] = ttt

    Juu = np.zeros((Nx*Ny, Nx*Ny))
    xJuu = np.zeros((Nx, Nx))
    yJuu = np.zeros((Ny, Ny))
    xJuu[:,0] = xJuu1
    yJuu[:,0] = yJuu1
    xyJuu = yJuu[:,0].reshape(-1,1) @ xJuu[:,0].reshape(1,-1)

    for jjj in range(1, Nx):
        xJuu[jjj:Nx,jjj] = xJuu1[0:Nx-jjj]
        xJuu[0:jjj+1,jjj] = xJuu1[jjj::-1]

    for jjj in range(1, Ny):
        yJuu[jjj:Ny,jjj] = yJuu1[0:Ny-jjj]
        yJuu[0:jjj+1,jjj] = yJuu1[jjj::-1]

    for j in range(Nx):
        for k in range(Ny):
            xyJuu = xJuu[:,j].reshape(-1,1) @ yJuu[:,k].reshape(1,-1)
            Juu[k*Nx+j,:] = xyJuu.reshape(1, Nx*Ny, order='F')

    return Juu

#Example:python.exe cpsdf.py 46 -119 acetone FRNK
if __name__ == '__main__':
    argvs = sys.argv

    delta = 2
    xmin = float(argvs[2]) - delta
    xmax = float(argvs[2]) + delta
    ymin = float(argvs[1]) - delta
    ymax = float(argvs[1]) + delta
    path = 'MetData'
    site = 'PROS'
    if len(argvs) > 4:
        site = argvs[4]
        if site is not None and os.path.exists('MetData_' + site):
            path = 'MetData_' + site
    start = time.time()
    filenameTrajConc = path + '/list.xlsx'

    directoryTraj = path
    filenameOutput =  'results.xlsx'
    corLengthScale = 1
    r = 0.25; 
    DTracking = 72
    Nx = 60
    Ny = 60
    NGx = 150
    NGy = 150
    tau=0 
    nFactor = 3
    
    mDataCell, nMeasurements = loadData(filenameTrajConc)
    meas = []
    for j in range(nMeasurements):
        meas.append(mDataCell[j][0])
    meas = np.array(meas)
    DurationTracking = DTracking
    varmeas = np.var(meas)
    lss2 = np.log(r*varmeas)
    lsf2 = np.log((1-r)*varmeas/DurationTracking/DurationTracking)
    ll2 = np.log(corLengthScale**2)
    dx = (xmax-xmin)/Nx
    dy = (ymax-ymin)/Ny
    NxPlus1 = Nx+1
    NyPlus1 = Ny+1
    xdlist = np.linspace(xmin,xmax,Nx+1)
    ydlist = np.linspace(ymin,ymax,Ny+1)
    xminlist = xdlist[0:Nx]
    xmaxlist = xdlist[1:Nx+1]
    xlist = (xminlist+xmaxlist)/2
    yminlist = ydlist[0:Ny]
    ymaxlist = ydlist[1:Ny+1]
    ylist = (yminlist+ymaxlist)/2
    xGList = np.linspace(xmin,xmax,NGx)
    yGList = np.linspace(ymin,ymax,NGy)
    totalmDataCellj2 = 0
    for j in range(nMeasurements):
        totalmDataCellj2 = totalmDataCellj2 + mDataCell[j][1]
    l = np.sqrt(np.exp(ll2))
    sf2 = np.exp(lsf2)
    ss2 = np.exp(lss2)
    xij = xlist.reshape(-1,1) @ np.ones(Ny).reshape(1,-1)
    yij = np.ones(Nx).reshape(-1,1) @ ylist.reshape(1,-1) 
    Wxu = np.zeros((nMeasurements, NGx*NGy))

    for j in range(nMeasurements):
        for k in range(mDataCell[j][1]):
            c = ReadDateFile(directoryTraj, mDataCell[j][2])
            ytr = c[:,9] 
            xtr = c[:,10]
            ttr = c[:,8]
            
            for kkk in range(nFactor):
                nPoints1 = len(xtr)
                xtr1 = np.zeros((nPoints1-1)*2+1)
                ytr1 = np.zeros((nPoints1-1)*2+1)
                ttr1 = np.zeros((nPoints1-1)*2+1)
                xtr1[0] = xtr[0]
                ytr1[0] = ytr[0]
                ttr1[0] = ttr[0]
                for kkkk in range(1, nPoints1):
                    xtr1[kkkk*2-1] = (xtr[kkkk-1]+xtr[kkkk])*.5
                    xtr1[kkkk*2] = xtr[kkkk]
                    ytr1[kkkk*2-1] = (ytr[kkkk-1]+ytr[kkkk])*.5
                    ytr1[kkkk*2] = ytr[kkkk]
                    ttr1[kkkk*2-1] = (ttr[kkkk-1]+ttr[kkkk])*.5
                    ttr1[kkkk*2] = ttr[kkkk]
                xtr = xtr1
                ytr = ytr1
                ttr = ttr1
            nPoints1 = len(xtr)
            xTrCent = (xtr[0:nPoints1-1] + xtr[1:nPoints1]) *.5
            yTrCent = (ytr[0:nPoints1-1] + ytr[1:nPoints1]) *.5
            dtCent = abs(ttr[0:nPoints1-1] - ttr[1:nPoints1])
            tCent = (ttr[0:nPoints1-1] + ttr[1:nPoints1]) *.5
    
            WxuSingleRow = createWxu(xGList,yGList,xTrCent,yTrCent,dtCent,tCent,tau)
            Wxu[j] += WxuSingleRow/mDataCell[j][1];
    
    WxuStar = np.zeros((Nx*Ny, NGx*NGy))
    Ncounter1 = 0
    for k in range(Ny):
        for j in range(Nx):
            WxuStarSingleRow = createWxuStar(xGList,yGList,xlist[j],ylist[k])
            Ncounter1 += 1
            WxuStar[Ncounter1-1] = WxuStarSingleRow

    WxuSp = sparse.coo_matrix(Wxu)
    del Wxu
    WxuStarSp = sparse.coo_matrix(WxuStar)
    del WxuStar

    l = np.sqrt(np.exp(ll2))
    sf2 = np.exp(lsf2)
    ss2 = np.exp(lss2)

    JuuSp = sparse.coo_matrix(createJuu(xGList, yGList, l))
    Kuu = sf2*JuuSp
    Kxx = WxuSp * Kuu * WxuSp.T
    A = Kxx + ss2 * np.eye(np.shape(Kxx)[0])
    KxStarx = WxuStarSp*Kuu*WxuSp.T
    fstarMean = KxStarx* np.linalg.solve(A, meas)
    psdfij = fstarMean.reshape(Nx, Ny, order='F')
    
    wb = load_workbook(filenameOutput)
    sheets = wb.worksheets
    sheet1 = sheets[0]
    for i in range(len(xlist)):
        sheet1.cell(1, i+2).value = xlist[i]
    for i in range(len(ylist)):
        sheet1.cell(i+2, 1).value = ylist[i]
    for i in range(len(xlist)):
        for j in range(len(ylist)):
            sheet1.cell(j+2, i+2).value = psdfij[i, j]
    wb.save(filenameOutput)
    gv2v1 = np.zeros((np.shape(psdfij)[0]+1, np.shape(psdfij)[1]+1))
    gv2v1[0:np.shape(psdfij)[0], 0:np.shape(psdfij)[1]] = psdfij
    
    # "1",PROS,46.39175309,-119.4117946, 110m
    # "15",FRNK,46.4172515,-119.2378376, 270m
    # 1	PROS	46.39175309	-119.4117946
    # 10	YAKB	46.57805411	-119.7261984
    # 11	300A	46.3641715	-119.2862936
    # 12	WYEB	46.48202169	-119.3912175
    # 13	100A	46.6892	-119.55069
    # 13	100N	46.68892672	-119.5508025
    # 14	WPPS	46.47007917	-119.3446059
    # 15	FRNK	46.4172515	-119.2378376
    # 16	GABL	46.59835954	-119.4602729
    # 17	RING	46.54494	-119.2375315
    # 18	RICH	46.3007828	-119.3010791
    # 19	PFP	46.54520946	-119.6325878
    # 2	EOC	46.39227355	-119.5370163
    # 20	RMTN	46.39481827	-119.5941445
    # 21	HMS	46.56280551	-119.5992092
    # 22	PASC	46.25718514	-119.1142853
    # 23	GABW	46.61178253	-119.5579632
    # 24	100F	46.63465653	-119.4523257
    # 25	VERN	46.6410943	-119.7274568
    # 26	BENT	46.29000253	-119.6076229
    # 27	VSTA	46.21834766	-119.2012749
    # 28	SURF	45.74361	-120.2175
    # 29	100K	46.65627226	-119.5797352
    # 3	ARMY	46.48925935	-119.55122
    # 30	HAMR	46.35636841	-119.3257338
    # 4	RSPG	46.50599944	-119.7000129
    # 5	EDNA	46.58711645	-119.3975151
    # 6	200E	46.55614	-119.5215505
    # 7	200W	46.5427357	-119.6626223
    # 8	BVLY	46.75236392	-119.9436184
    # 9	FFTF	46.42954177	-119.3597285
    mycolor = 'Navy'    
    # Measurements site location
    distanceLat = 0.95
    distanceLng = 0.15
    mylat = 46.39175309
    mylon = -119.4117946
    if site == 'PROS':
        site = 'PROS_1 (110m)'
    elif site == '300A': 
        site='300A_11' 
        mylat = 46.3641715
        mylon = -119.2862936
    elif site == 'WYEB': 
        site='WYEB_12' 
        mylat = 46.48202169
        mylon = -119.3912175
    elif site == '100A': 
        site='100A_13' 
        mylat = 46.6892
        mylon = -119.55069
    elif site == '100N': 
        site='100N_13' 
        mylat = 46.68892672
        mylon = -119.5508025
    elif site == 'WPPS': 
        site='WPPS_14' 
        mylat = 46.47007917
        mylon = -119.3446059
    elif site == 'FRNK':
        site = 'FRNK_15 (270m)'
        mycolor = 'Orange'
        mylat = 46.4172515
        mylon = -119.2378376
    elif site == 'GABL': 
        site='GABL_16' 
        mylat = 46.59835954
        mylon = -119.4602729
    elif site == 'RING': 
        site='RING_17' 
        mylat = 46.54494
        mylon = -119.2375315
    elif site == 'RICH': 
        site='RICH_18' 
        mylat = 46.3007828
        mylon = -119.3010791
    elif site == 'PFP': 
        site='PFP_19' 
        mylat = 46.54520946
        mylon = -119.6325878
    elif site == 'EOC': 
        site='EOC_2' 
        mylat = 46.39227355
        mylon = -119.5370163
    elif site == 'RMTN': 
        site='RMTN_20' 
        mylat = 46.39481827
        mylon = -119.5941445
    elif site == 'HMS': 
        site='HMS_21' 
        mylat = 46.56280551
        mylon = -119.5992092
    elif site == 'PASC': 
        site='PASC_22' 
        mylat = 46.25718514
        mylon = -119.1142853
    elif site == 'GABW': 
        site='GABW_23' 
        mylat = 46.61178253
        mylon = -119.5579632
    elif site == '100F': 
        site='100F_24' 
        mylat = 46.63465653
        mylon = -119.4523257
    elif site == 'VERN': 
        site='VERN_25' 
        mylat = 46.6410943
        mylon = -119.7274568
    elif site == 'BENT': 
        site='BENT_26' 
        mylat = 46.29000253
        mylon = -119.6076229
    elif site == 'VSTA': 
        site='VSTA_27' 
        mylat = 46.21834766
        mylon = -119.2012749
    elif site == 'SURF': 
        site='SURF_28' 
        mylat = 45.74361
        mylon = -120.2175
    elif site == '100K': 
        site='100K_29' 
        mylat = 46.65627226
        mylon = -119.5797352
    elif site == 'ARMY': 
        site='ARMY_3' 
        mylat = 46.48925935
        mylon = -119.55122
    elif site == 'HAMR': 
        site='HAMR_30' 
        mylat = 46.35636841
        mylon = -119.3257338
    elif site == 'RSPG': 
        site='RSPG_4' 
        mylat = 46.50599944
        mylon = -119.7000129
    elif site == 'EDNA': 
        site='EDNA_5' 
        mylat = 46.58711645
        mylon = -119.3975151
    elif site == '200E': 
        site='200E_6' 
        mylat = 46.55614
        mylon = -119.5215505
    elif site == '200W': 
        site='200W_7' 
        mylat = 46.5427357
        mylon = -119.6626223
    elif site == 'BVLY': 
        site='BVLY_8' 
        mylat = 46.75236392
        mylon = -119.9436184
    elif site == 'FFTF': 
        site='FFTF_9' 
        mylat = 46.42954177
        mylon = -119.3597285
    lons,lats= np.meshgrid([ x + distanceLng for x in xdlist],[ y + distanceLat for y in ydlist]) 

    fig = plt.figure(figsize = (7, 6))
    offset = 2
    lclon=lons.min() - offset
    rclon=lons.max() + offset
    lclat=lats.min() - offset
    rclat=lats.max() + offset
    clon = lons.mean()
    clat = lats.mean()
    mymap = Basemap(projection='mill',llcrnrlon=lclon, llcrnrlat=lclat, urcrnrlon=rclon,urcrnrlat=rclat, lon_0=lclon, lat_0=clat, resolution='h')
    mymap.shadedrelief()
    x,y = mymap(lons,lats)
    mymap.drawparallels(np.arange(lclat, rclat,1),labels=[1,0,0,0], fontsize=6)
    mymap.drawmeridians(np.arange(mymap.lonmin,mymap.lonmax,1),labels=[0,0,0,1], fontsize=6)
    mapcontourf = mymap.contourf(x,y,gv2v1)
    cb = mymap.colorbar(mapcontourf, size="5%")
    cb.ax.tick_params(labelsize=6) 
    cb.set_label('Probability', fontsize=12)
        
    x,y = mymap(-119.4117946,46.39175309)
    plt.plot(x,y,'ok', markersize=2, color='Navy')
    plt.text(x,y, '1', fontsize=6, color='Navy')
    x,y = mymap(-119.2378376,46.4172515)
    plt.plot(x,y,'ok', markersize=2, color='Orange')
    plt.text(x,y, '15', fontsize=6, color='Orange')
    x,y = mymap(-119.2752,46.2804)
    plt.plot(x,y,'ok', markersize=2)
    plt.text(x,y, ' Richland', fontsize=8)
    x,y = mymap(-122.3321,47.6062)
    plt.plot(x,y,'ok', markersize=2)
    plt.text(x,y, 'Seattle', fontsize=8)
    

    plt.title(site, fontsize=20, color=mycolor)
    filename = argvs[3] + '_' + site + ".png"
    plt.savefig(filename, dpi=300, format='png')
    plt.show()